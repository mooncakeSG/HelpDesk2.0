#!/usr/bin/env python3
"""
Export all tickets from database to Excel daily log file with proper formatting
"""

import pandas as pd
import requests
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load environment variables
load_dotenv()

# SQLiteCloud configuration
API_KEY = os.getenv('SQLITECLOUD_API_KEY')
API_URL = os.getenv('SQLITECLOUD_URL')

headers = {
    'Authorization': f'Bearer {API_KEY}',
    'Content-Type': 'application/json'
}

def execute_query(query):
    """Execute a SQL query on SQLiteCloud"""
    try:
        payload = {"sql": query}
        response = requests.post(API_URL, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if 'data' in result:
                return result['data']
            return result
        else:
            print(f"❌ Query failed: {response.text}")
            return None
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

def get_all_tickets():
    """Get all tickets from database"""
    query = '''
        USE DATABASE 'my-database';
        SELECT 
            id as TicketID,
            timestamp as DateOpened,
            name as ReporterName,
            email as ReporterContact,
            assigned_agent as AssignedAgent,
            priority as Priority,
            status as Status,
            issue as FullDescription,
            notes as ResolutionNotes
        FROM tickets 
        ORDER BY id
    '''
    
    tickets = execute_query(query)
    return tickets

def get_kb_article(issue):
    """Determine KB article based on issue type"""
    issue_lower = str(issue).lower()
    
    if 'password' in issue_lower and 'forgotten' in issue_lower:
        return 'KB_Password_Reset'
    elif 'lockout' in issue_lower or 'locked' in issue_lower:
        return 'KB_Password_Reset'
    elif 'disabled' in issue_lower:
        return 'KB_Account_Enable'
    elif 'outlook' in issue_lower or 'authentication' in issue_lower:
        return 'KB_MFA_Reset'
    elif 'mfa' in issue_lower and 'lost' in issue_lower:
        return 'KB_MFA_Reset'
    elif 'temporary' in issue_lower and 'contractor' in issue_lower:
        return 'KB_Temp_Account'
    elif 'suspicious' in issue_lower or 'unknown' in issue_lower:
        return 'KB_Security_Check'
    elif 'expired' in issue_lower:
        return 'KB_Password_Reset'
    else:
        return 'KB_General_Support'

def create_daily_log_excel(tickets):
    """Create Excel file with daily log of all tickets with proper formatting"""
    
    # Get current date for filename
    current_date = datetime.now().strftime("%Y%m%d")
    filename = f"IT_Helpdesk_Daily_Log_Formatted_{current_date}.xlsx"
    
    print(f"📊 Creating formatted daily log: {filename}")
    
    # Create DataFrame
    df = pd.DataFrame(tickets)
    
    # Add calculated columns
    df['BriefSummary'] = df['FullDescription'].str[:80] + "..."  # Longer summary
    df['IncidentCategory'] = 'Account / Authentication'
    df['KBArticleLinked'] = df['FullDescription'].apply(get_kb_article)
    df['ResolutionDate'] = df['DateOpened']
    df['TimeToResolve'] = 'Same Day'
    
    # Reorder columns to match your original Excel format
    column_order = [
        'TicketID', 'DateOpened', 'ReporterName', 'ReporterContact', 
        'AssignedAgent', 'IncidentCategory', 'Priority', 'BriefSummary', 
        'FullDescription', 'Status', 'ResolutionNotes', 'KBArticleLinked',
        'ResolutionDate', 'TimeToResolve'
    ]
    
    df = df[column_order]
    
    # Create Excel file with proper formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main daily log sheet
        df.to_excel(writer, sheet_name='Daily Log', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Daily Log']
        
        # Apply formatting
        format_excel_sheet(workbook, worksheet, df)
        
        # Create summary statistics
        create_summary_sheet(workbook, df)
        
        # Create agent workload sheet
        create_agent_workload_sheet(workbook, df)
    
    print(f"✅ Formatted daily log created successfully: {filename}")
    return filename

def format_excel_sheet(workbook, worksheet, df):
    """Apply comprehensive formatting to the Excel sheet"""
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Apply data formatting and set column widths
    column_widths = {
        'A': 8,   # TicketID
        'B': 12,  # DateOpened
        'C': 20,  # ReporterName
        'D': 25,  # ReporterContact
        'E': 18,  # AssignedAgent
        'F': 20,  # IncidentCategory
        'G': 10,  # Priority
        'H': 30,  # BriefSummary
        'I': 50,  # FullDescription
        'J': 12,  # Status
        'K': 60,  # ResolutionNotes
        'L': 18,  # KBArticleLinked
        'M': 12,  # ResolutionDate
        'N': 15   # TimeToResolve
    }
    
    # Set column widths
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Apply formatting to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Set row heights for better readability
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 60  # Increased row height
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

def create_summary_sheet(workbook, df):
    """Create a summary statistics sheet"""
    
    summary_data = {
        'Metric': [
            'Total Tickets',
            'Resolved Tickets',
            'Open Tickets',
            'In Progress Tickets',
            'High Priority Tickets',
            'Medium Priority Tickets',
            'Low Priority Tickets',
            'Tickets by Azola Xabadiya',
            'Tickets by Keawin Koesnel',
            'Tickets by System Admin',
            'Unassigned Tickets'
        ],
        'Count': [
            len(df),
            len(df[df['Status'] == 'Resolved']),
            len(df[df['Status'] == 'Open']),
            len(df[df['Status'] == 'In Progress']),
            len(df[df['Priority'] == 'High']),
            len(df[df['Priority'] == 'Medium']),
            len(df[df['Priority'] == 'Low']),
            len(df[df['AssignedAgent'] == 'Azola Xabadiya']),
            len(df[df['AssignedAgent'] == 'Keawin Koesnel']),
            len(df[df['AssignedAgent'] == 'System Admin']),
            len(df[df['AssignedAgent'] == 'Unassigned'])
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet('Summary')
    
    # Add data to summary sheet
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        summary_sheet.append(r)
    
    # Format summary sheet
    format_summary_sheet(summary_sheet)

def create_agent_workload_sheet(workbook, df):
    """Create an agent workload sheet"""
    
    agent_workload = df.groupby('AssignedAgent').agg({
        'TicketID': 'count',
        'Priority': lambda x: (x == 'High').sum(),
        'Status': lambda x: (x == 'Resolved').sum()
    }).rename(columns={
        'TicketID': 'Total_Tickets',
        'Priority': 'High_Priority',
        'Status': 'Resolved_Tickets'
    })
    
    # Create agent workload sheet
    workload_sheet = workbook.create_sheet('Agent Workload')
    
    # Add data to workload sheet
    for r in dataframe_to_rows(agent_workload, index=True, header=True):
        workload_sheet.append(r)
    
    # Format workload sheet
    format_workload_sheet(workload_sheet)

def format_summary_sheet(worksheet):
    """Format the summary sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 15

def format_workload_sheet(worksheet):
    """Format the agent workload sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15

def main():
    print("📊 IT Helpdesk Formatted Daily Log Export Tool")
    print("==============================================")
    print()
    
    # Get all tickets
    print("🔍 Retrieving all tickets from database...")
    tickets = get_all_tickets()
    
    if not tickets:
        print("❌ No tickets found in database")
        return
    
    print(f"📋 Found {len(tickets)} tickets to export")
    print()
    
    # Create Excel file
    filename = create_daily_log_excel(tickets)
    
    print()
    print("📈 Export Summary:")
    print(f"   📄 File created: {filename}")
    print(f"   🎫 Total tickets: {len(tickets)}")
    print(f"   📊 Sheets included: Daily Log, Summary, Agent Workload")
    print(f"   🎨 Formatting: Headers, borders, proper column widths, row heights")
    
    print()
    print("✅ Formatted daily log export completed successfully!")
    print(f"📁 File location: {os.path.abspath(filename)}")
    print()
    print("📋 The Excel file includes:")
    print("   • Daily Log sheet with proper column widths and formatting")
    print("   • Summary sheet with statistics")
    print("   • Agent Workload sheet with performance metrics")
    print("   • Frozen header row for easy scrolling")
    print("   • Proper text wrapping and row heights")
    print("   • Professional styling with borders and colors")

if __name__ == "__main__":
    main()
