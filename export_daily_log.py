#!/usr/bin/env python3
"""
Export all tickets from database to Excel daily log file
"""

import pandas as pd
import requests
import os
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# SQLiteCloud configuration
API_KEY = os.getenv('SQLITECLOUD_API_KEY')
API_URL = os.getenv('SQLITECLOUD_URL')

headers = {
    'Authorization': f'Bearer {API_KEY}',
    'Content-Type': 'application/json'
}

def execute_query(query):
    """Execute a SQL query on SQLiteCloud"""
    try:
        payload = {"sql": query}
        response = requests.post(API_URL, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            if 'data' in result:
                return result['data']
            return result
        else:
            print(f"❌ Query failed: {response.text}")
            return None
    except Exception as e:
        print(f"❌ Error: {e}")
        return None

def get_all_tickets():
    """Get all tickets from database"""
    query = '''
        USE DATABASE 'my-database';
        SELECT 
            id as TicketID,
            timestamp as DateOpened,
            name as ReporterName,
            email as ReporterContact,
            assigned_agent as AssignedAgent,
            priority as Priority,
            status as Status,
            issue as FullDescription,
            notes as ResolutionNotes
        FROM tickets 
        ORDER BY id
    '''
    
    tickets = execute_query(query)
    return tickets

def create_daily_log_excel(tickets):
    """Create Excel file with daily log of all tickets"""
    
    # Get current date for filename
    current_date = datetime.now().strftime("%Y%m%d")
    filename = f"IT_Helpdesk_Daily_Log_{current_date}.xlsx"
    
    print(f"📊 Creating daily log: {filename}")
    
    # Create DataFrame
    df = pd.DataFrame(tickets)
    
    # Add calculated columns
    df['BriefSummary'] = df['FullDescription'].str[:50] + "..."
    df['IncidentCategory'] = 'Account / Authentication'  # Based on your Excel data
    df['KBArticleLinked'] = df.apply(get_kb_article, axis=1)
    df['ResolutionDate'] = df['DateOpened']  # Assuming resolved on same day
    df['TimeToResolve'] = 'Same Day'
    
    # Reorder columns to match your original Excel format
    column_order = [
        'TicketID', 'DateOpened', 'ReporterName', 'ReporterContact', 
        'AssignedAgent', 'IncidentCategory', 'Priority', 'BriefSummary', 
        'FullDescription', 'Status', 'ResolutionNotes', 'KBArticleLinked',
        'ResolutionDate', 'TimeToResolve'
    ]
    
    df = df[column_order]
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Daily Log', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Daily Log']
        
        # Apply formatting
        format_excel_sheet(workbook, worksheet, df)
        
        # Create summary sheet
        create_summary_sheet(workbook, df)
    
    print(f"✅ Daily log created successfully: {filename}")
    return filename

def get_kb_article(row):
    """Determine KB article based on issue type"""
    issue = str(row['FullDescription']).lower()
    
    if 'password' in issue and 'forgotten' in issue:
        return 'KB_Password_Reset'
    elif 'lockout' in issue or 'locked' in issue:
        return 'KB_Password_Reset'
    elif 'disabled' in issue:
        return 'KB_Account_Enable'
    elif 'outlook' in issue or 'authentication' in issue:
        return 'KB_MFA_Reset'
    elif 'mfa' in issue and 'lost' in issue:
        return 'KB_MFA_Reset'
    elif 'temporary' in issue and 'contractor' in issue:
        return 'KB_Temp_Account'
    elif 'suspicious' in issue or 'unknown' in issue:
        return 'KB_Security_Check'
    elif 'expired' in issue:
        return 'KB_Password_Reset'
    else:
        return 'KB_General_Support'

def format_excel_sheet(workbook, worksheet, df):
    """Apply formatting to the Excel sheet"""
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders to all cells
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

def create_summary_sheet(workbook, df):
    """Create a summary statistics sheet"""
    
    # Create summary data
    summary_data = {
        'Metric': [
            'Total Tickets',
            'Resolved Tickets',
            'Open Tickets',
            'In Progress Tickets',
            'High Priority Tickets',
            'Medium Priority Tickets',
            'Low Priority Tickets',
            'Tickets by Azola Xabadiya',
            'Tickets by Keawin Koesnel',
            'Tickets by System Admin',
            'Unassigned Tickets'
        ],
        'Count': [
            len(df),
            len(df[df['Status'] == 'Resolved']),
            len(df[df['Status'] == 'Open']),
            len(df[df['Status'] == 'In Progress']),
            len(df[df['Priority'] == 'High']),
            len(df[df['Priority'] == 'Medium']),
            len(df[df['Priority'] == 'Low']),
            len(df[df['AssignedAgent'] == 'Azola Xabadiya']),
            len(df[df['AssignedAgent'] == 'Keawin Koesnel']),
            len(df[df['AssignedAgent'] == 'System Admin']),
            len(df[df['AssignedAgent'] == 'Unassigned'])
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    
    # Write summary to new sheet
    with pd.ExcelWriter('temp_summary.xlsx', engine='openpyxl') as temp_writer:
        summary_df.to_excel(temp_writer, sheet_name='Summary', index=False)
        temp_workbook = temp_writer.book
        temp_worksheet = temp_writer.sheets['Summary']
    
    # Copy summary sheet to main workbook
    summary_sheet = workbook.create_sheet('Summary')
    
    # Copy data
    for row in temp_worksheet.iter_rows():
        for cell in row:
            new_cell = summary_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font
                new_cell.fill = cell.fill
                new_cell.border = cell.border
                new_cell.alignment = cell.alignment
    
    # Clean up temp file
    os.remove('temp_summary.xlsx')

def main():
    print("📊 IT Helpdesk Daily Log Export Tool")
    print("====================================")
    print()
    
    # Get all tickets
    print("🔍 Retrieving all tickets from database...")
    tickets = get_all_tickets()
    
    if not tickets:
        print("❌ No tickets found in database")
        return
    
    print(f"📋 Found {len(tickets)} tickets to export")
    print()
    
    # Create Excel file
    filename = create_daily_log_excel(tickets)
    
    print()
    print("📈 Export Summary:")
    print(f"   📄 File created: {filename}")
    print(f"   🎫 Total tickets: {len(tickets)}")
    print(f"   📊 Sheets included: Daily Log, Summary")
    print(f"   🎨 Formatting: Headers, borders, auto-width")
    
    print()
    print("✅ Daily log export completed successfully!")
    print(f"📁 File location: {os.path.abspath(filename)}")
    print()
    print("📋 The Excel file includes:")
    print("   • All ticket details with proper formatting")
    print("   • Summary statistics sheet")
    print("   • KB article references")
    print("   • Resolution notes and steps")
    print("   • Agent assignments and priorities")

if __name__ == "__main__":
    main()
