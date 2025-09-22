#!/usr/bin/env python3
"""
Create comprehensive Excel log with all ticket information from database including Week 2 tickets
"""

import pandas as pd
import requests
import os
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load environment variables
load_dotenv()

# SQLiteCloud configuration
API_KEY = os.getenv('SQLITECLOUD_API_KEY')
API_URL = os.getenv('SQLITECLOUD_URL')

if not API_KEY or not API_URL:
    print("❌ Error: SQLITECLOUD_API_KEY and SQLITECLOUD_URL must be set in environment variables")
    exit(1)

headers = {
    'Authorization': f'Bearer {API_KEY}',
    'Content-Type': 'application/json'
}

def execute_query(query, params=None):
    """Execute a SQL query on SQLiteCloud"""
    try:
        payload = {"sql": query}
        response = requests.post(API_URL, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            return result
        else:
            print(f"❌ Query failed with status code: {response.status_code}")
            return None
    except Exception as e:
        print(f"❌ Error executing query: {e}")
        return None

def fetch_all_tickets():
    """Fetch all tickets from the database"""
    print("📊 Fetching all tickets from database...")
    
    query = '''
        USE DATABASE 'my-database';
        SELECT * FROM tickets ORDER BY id
    '''
    
    result = execute_query(query)
    if result and 'data' in result:
        return result['data']
    else:
        print("❌ Failed to fetch tickets from database")
        return []

def create_comprehensive_excel():
    """Create comprehensive Excel file with all ticket information"""
    
    # Get current date for filename
    current_date = datetime.now().strftime("%Y%m%d")
    filename = f"IT_Helpdesk_Comprehensive_Log_{current_date}.xlsx"
    
    print(f"📊 Creating comprehensive log: {filename}")
    
    # Fetch all tickets from database
    tickets_data = fetch_all_tickets()
    
    if not tickets_data:
        print("❌ No tickets found in database")
        return None
    
    print(f"✅ Found {len(tickets_data)} tickets in database")
    
    # Convert database tickets to comprehensive format
    comprehensive_tickets = []
    
    for ticket in tickets_data:
        # Determine incident category based on issue content
        issue = ticket.get('issue', '').lower()
        if any(keyword in issue for keyword in ['password', 'account', 'login', 'authentication', 'mfa', 'lockout']):
            incident_category = 'Account / Authentication'
        elif any(keyword in issue for keyword in ['laptop', 'hardware', 'bsod', 'fan', 'usb', 'monitor']):
            incident_category = 'Hardware Support'
        elif any(keyword in issue for keyword in ['network', 'wifi', 'ethernet', 'internet', 'connection']):
            incident_category = 'Network Support'
        elif any(keyword in issue for keyword in ['software', 'windows', 'browser', 'outlook', 'application']):
            incident_category = 'Software Support'
        elif any(keyword in issue for keyword in ['stolen', 'security', 'burglary', 'remote wipe']):
            incident_category = 'Security Incident'
        else:
            incident_category = 'General Support'
        
        # Generate agent response based on issue type
        agent_response = generate_agent_response(ticket.get('issue', ''), ticket.get('priority', 'Medium'))
        
        comprehensive_ticket = {
            'TicketID': ticket.get('id', ''),
            'DateOpened': ticket.get('timestamp', ''),
            'ReporterName': ticket.get('name', ''),
            'ReporterContact': ticket.get('email', ''),
            'AssignedAgent': ticket.get('assigned_agent', 'Unassigned'),
            'IncidentCategory': incident_category,
            'Priority': ticket.get('priority', 'Medium'),
            'BriefSummary': ticket.get('issue', ''),
            'Status': ticket.get('status', 'Open'),
            'ResolutionNotes': ticket.get('notes', ''),
            'KBArticleLinked': generate_kb_article(incident_category),
            'ResolutionDate': ticket.get('timestamp', '') if ticket.get('status') == 'Resolved' else '',
            'TimeToResolve': 'Same Day' if ticket.get('status') == 'Resolved' else 'Pending',
            'AgentResponse': agent_response,
            'Category': ticket.get('category', 'General')
        }
        
        comprehensive_tickets.append(comprehensive_ticket)
    
    # Create DataFrame
    df = pd.DataFrame(comprehensive_tickets)
    
    # Create Excel file with proper formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main comprehensive log sheet
        df.to_excel(writer, sheet_name='Comprehensive Log', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Comprehensive Log']
        
        # Apply formatting
        format_comprehensive_sheet(workbook, worksheet, df)
        
        # Create summary statistics
        create_summary_sheet(workbook, df)
        
        # Create agent workload sheet
        create_agent_workload_sheet(workbook, df)
        
        # Create conversation log sheet
        create_conversation_log_sheet(workbook, comprehensive_tickets)
        
        # Create category breakdown sheet
        create_category_breakdown_sheet(workbook, df)
    
    print(f"✅ Comprehensive log created successfully: {filename}")
    return filename

def generate_agent_response(issue, priority):
    """Generate appropriate agent response based on issue type"""
    issue_lower = issue.lower()
    
    if 'password' in issue_lower or 'lockout' in issue_lower:
        return "Hello, I can help with that. For security, I need to verify your identity first.\n\nPlease provide your full name and username, and I will call you back on the number we have on file to confirm and then reset your password.\n\nA temporary password will be set, and you'll be required to change it on first login."
    
    elif 'stolen' in issue_lower or 'burglary' in issue_lower:
        return "I understand this is urgent. I'm immediately initiating our security protocol.\n\nI'll start the remote wipe procedure and coordinate with security to ensure all data is protected. A replacement device will be arranged as quickly as possible.\n\nPlease stay on the line while I handle this security incident."
    
    elif 'bsod' in issue_lower or 'crash' in issue_lower:
        return "I can help diagnose this system crash. Let me gather some information first.\n\nCan you tell me what you were doing when the crash occurred? Also, please note any error messages you saw.\n\nI'll run some diagnostic checks and get this resolved for you."
    
    elif 'network' in issue_lower or 'wifi' in issue_lower or 'internet' in issue_lower:
        return "I'll help you get your network connection working. Let me check a few things.\n\nFirst, let's verify your network settings and test connectivity. I'll also check if this is affecting other users.\n\nThis should be a quick fix once I identify the issue."
    
    elif 'hardware' in issue_lower or 'laptop' in issue_lower or 'fan' in issue_lower:
        return "I can help with this hardware issue. Let me assess the situation.\n\nI'll check system diagnostics and determine if this requires immediate replacement or if we can resolve it with troubleshooting.\n\nI'll keep you updated on the progress and timeline."
    
    else:
        return "Thank you for contacting IT support. I'm here to help resolve this issue.\n\nLet me gather some additional information and then I'll work on getting this sorted out for you.\n\nI'll keep you informed of my progress and any next steps."

def generate_kb_article(category):
    """Generate appropriate KB article reference based on category"""
    kb_articles = {
        'Account / Authentication': 'KB_Account_Management',
        'Hardware Support': 'KB_Hardware_Troubleshooting',
        'Network Support': 'KB_Network_Connectivity',
        'Software Support': 'KB_Software_Support',
        'Security Incident': 'KB_Security_Procedures',
        'General Support': 'KB_General_Support'
    }
    return kb_articles.get(category, 'KB_General_Support')

def format_comprehensive_sheet(workbook, worksheet, df):
    """Apply comprehensive formatting to the Excel sheet"""
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths for comprehensive data
    column_widths = {
        'A': 8,   # TicketID
        'B': 18,  # DateOpened
        'C': 20,  # ReporterName
        'D': 30,  # ReporterContact
        'E': 18,  # AssignedAgent
        'F': 20,  # IncidentCategory
        'G': 10,  # Priority
        'H': 60,  # BriefSummary
        'I': 12,  # Status
        'J': 80,  # ResolutionNotes
        'K': 18,  # KBArticleLinked
        'L': 18,  # ResolutionDate
        'M': 15,  # TimeToResolve
        'N': 60,  # AgentResponse
        'O': 25   # Category
    }
    
    # Set column widths
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Apply formatting to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Set row heights for better readability
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 80
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

def create_summary_sheet(workbook, df):
    """Create a summary statistics sheet"""
    
    # Calculate statistics
    total_tickets = len(df)
    resolved_tickets = len(df[df['Status'] == 'Resolved'])
    open_tickets = len(df[df['Status'] == 'Open'])
    in_progress_tickets = len(df[df['Status'] == 'In Progress'])
    
    high_priority = len(df[df['Priority'] == 'High'])
    medium_priority = len(df[df['Priority'] == 'Medium'])
    low_priority = len(df[df['Priority'] == 'Low'])
    
    # Agent statistics
    agent_stats = df['AssignedAgent'].value_counts()
    
    # Category statistics
    category_stats = df['Category'].value_counts()
    
    summary_data = {
        'Metric': [
            'Total Tickets',
            'Resolved Tickets',
            'Open Tickets',
            'In Progress Tickets',
            'High Priority Tickets',
            'Medium Priority Tickets',
            'Low Priority Tickets',
            '',
            'Tickets by Azola Xabadiya',
            'Tickets by Keawin Koesnel',
            'Tickets by System Admin',
            'Unassigned Tickets',
            '',
            'Week 1: Account & Communications',
            'Week 2: Software & Hardware'
        ],
        'Count': [
            total_tickets,
            resolved_tickets,
            open_tickets,
            in_progress_tickets,
            high_priority,
            medium_priority,
            low_priority,
            '',
            agent_stats.get('Azola Xabadiya', 0),
            agent_stats.get('Keawin Koesnel', 0),
            agent_stats.get('System Admin', 0),
            agent_stats.get('Unassigned', 0),
            '',
            category_stats.get('Week 1: Account and Communications Support', 0),
            category_stats.get('Week 2: Software & Hardware Support', 0)
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    
    # Create summary sheet
    summary_sheet = workbook.create_sheet('Summary')
    
    # Add data to summary sheet
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        summary_sheet.append(r)
    
    # Format summary sheet
    format_summary_sheet(summary_sheet)

def create_agent_workload_sheet(workbook, df):
    """Create an agent workload sheet"""
    
    agent_workload = df.groupby('AssignedAgent').agg({
        'TicketID': 'count',
        'Priority': lambda x: (x == 'High').sum(),
        'Status': lambda x: (x == 'Resolved').sum()
    }).rename(columns={
        'TicketID': 'Total_Tickets',
        'Priority': 'High_Priority',
        'Status': 'Resolved_Tickets'
    })
    
    # Create agent workload sheet
    workload_sheet = workbook.create_sheet('Agent Workload')
    
    # Add data to workload sheet
    for r in dataframe_to_rows(agent_workload, index=True, header=True):
        workload_sheet.append(r)
    
    # Format workload sheet
    format_workload_sheet(workload_sheet)

def create_conversation_log_sheet(workbook, tickets):
    """Create a conversation log sheet showing user issues and agent responses"""
    
    conversation_sheet = workbook.create_sheet('Conversation Log')
    
    # Add headers
    headers = ['Timestamp', 'Ticket ID', 'Category', 'User Issue', 'Agent Response']
    conversation_sheet.append(headers)
    
    # Add conversation data
    for ticket in tickets:
        conversation_sheet.append([
            ticket['DateOpened'],
            ticket['TicketID'],
            ticket['Category'],
            ticket['BriefSummary'],
            ticket['AgentResponse']
        ])
    
    # Format conversation sheet
    format_conversation_sheet(conversation_sheet)

def create_category_breakdown_sheet(workbook, df):
    """Create a category breakdown sheet"""
    
    category_sheet = workbook.create_sheet('Category Breakdown')
    
    # Group by category and calculate statistics
    category_breakdown = df.groupby('Category').agg({
        'TicketID': 'count',
        'Priority': lambda x: (x == 'High').sum(),
        'Status': lambda x: (x == 'Resolved').sum()
    }).rename(columns={
        'TicketID': 'Total_Tickets',
        'Priority': 'High_Priority',
        'Status': 'Resolved_Tickets'
    })
    
    # Add data to category sheet
    for r in dataframe_to_rows(category_breakdown, index=True, header=True):
        category_sheet.append(r)
    
    # Format category sheet
    format_category_sheet(category_sheet)

def format_summary_sheet(worksheet):
    """Format the summary sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 15

def format_workload_sheet(worksheet):
    """Format the agent workload sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15

def format_conversation_sheet(worksheet):
    """Format the conversation log sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 18  # Timestamp
    worksheet.column_dimensions['B'].width = 10  # Ticket ID
    worksheet.column_dimensions['C'].width = 25  # Category
    worksheet.column_dimensions['D'].width = 80  # User Issue
    worksheet.column_dimensions['E'].width = 80  # Agent Response
    
    # Set row heights
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 60

def format_category_sheet(worksheet):
    """Format the category breakdown sheet"""
    
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 35  # Category
    worksheet.column_dimensions['B'].width = 15  # Total Tickets
    worksheet.column_dimensions['C'].width = 15  # High Priority
    worksheet.column_dimensions['D'].width = 15  # Resolved Tickets

def main():
    print("📊 IT Helpdesk Comprehensive Log Creator (Updated)")
    print("=================================================")
    print()
    
    # Create Excel file
    filename = create_comprehensive_excel()
    
    if filename:
        print()
        print("📈 Export Summary:")
        print(f"   📄 File created: {filename}")
        print(f"   🎫 Total tickets: {len(fetch_all_tickets())}")
        print(f"   📊 Sheets included: Comprehensive Log, Summary, Agent Workload, Conversation Log, Category Breakdown")
        print(f"   🎨 Formatting: Headers, borders, proper column widths, row heights")
        
        print()
        print("✅ Comprehensive log export completed successfully!")
        print(f"📁 File location: {os.path.abspath(filename)}")
        print()
        print("📋 The Excel file includes:")
        print("   • Comprehensive Log sheet with all ticket details and agent responses")
        print("   • Summary sheet with statistics including Week 1 and Week 2 breakdown")
        print("   • Agent Workload sheet with performance metrics")
        print("   • Conversation Log sheet showing user issues and agent responses")
        print("   • Category Breakdown sheet showing Week 1 vs Week 2 statistics")
        print("   • Professional styling with borders and colors")
        print("   • Proper text wrapping and row heights for readability")
    else:
        print("❌ Failed to create comprehensive log")

if __name__ == "__main__":
    main()
