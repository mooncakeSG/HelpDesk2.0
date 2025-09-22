#!/usr/bin/env python3
"""
Create Week 1 Knowledge Base from IT Helpdesk Comprehensive Log
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_week1_knowledge_base():
    """Create Week 1 Knowledge Base Excel file"""
    
    # Get current date for filename
    current_date = datetime.now().strftime("%Y%m%d")
    filename = f"Week1_Knowledge_Base_{current_date}.xlsx"
    
    print(f"📚 Creating Week 1 Knowledge Base: {filename}")
    
    # Create Excel file with multiple knowledge base sheets
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Create different knowledge base sections
        create_incident_patterns_sheet(writer)
        create_solution_playbook_sheet(writer)
        create_agent_performance_sheet(writer)
        create_common_issues_sheet(writer)
        create_escalation_procedures_sheet(writer)
        create_week1_summary_sheet(writer)
    
    print(f"✅ Week 1 Knowledge Base created successfully: {filename}")
    return filename

def create_incident_patterns_sheet(writer):
    """Create incident patterns analysis sheet"""
    
    # Incident patterns data
    patterns_data = {
        'Incident Type': [
            'Password Reset Requests',
            'Account Lockouts',
            'Recurring Lockouts',
            'Account Disabled',
            'Outlook Authentication',
            'MFA Device Issues',
            'Password Expiration',
            'Temporary Access',
            'Security Incidents'
        ],
        'Frequency': [4, 3, 1, 2, 1, 1, 1, 1, 1],
        'Priority Level': ['Medium', 'Medium', 'Medium', 'Medium', 'Low', 'High', 'Medium', 'Low', 'High'],
        'Average Resolution Time': ['Same Day', 'Same Day', 'Same Day', 'Same Day', 'Same Day', 'Same Day', 'Same Day', 'Same Day', 'Same Day'],
        'Common Root Causes': [
            'User forgot password after holiday/break',
            'Failed login attempts (3+ attempts)',
            'Cached credentials on multiple devices',
            'Account disabled by mistake or policy',
            'Cached Office 365 credentials',
            'Lost/stolen mobile device',
            'Password policy expiration',
            'Contractor access requirements',
            'Suspicious login attempts from unknown locations'
        ],
        'Prevention Strategies': [
            'Password reminder notifications before expiration',
            'User education on correct password entry',
            'Regular credential cache cleanup',
            'Review account disablement policies',
            'Regular Office 365 credential refresh',
            'Backup MFA device setup',
            'Proactive password expiration notifications',
            'Standardized contractor onboarding process',
            'Enhanced security monitoring and alerts'
        ]
    }
    
    patterns_df = pd.DataFrame(patterns_data)
    patterns_df.to_excel(writer, sheet_name='Incident Patterns', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Incident Patterns']
    format_knowledge_sheet(workbook, worksheet, 'Incident Patterns Analysis')

def create_solution_playbook_sheet(writer):
    """Create solution playbook sheet"""
    
    playbook_data = {
        'Issue Category': [
            'Password Reset',
            'Password Reset',
            'Account Unlock',
            'Account Unlock',
            'Recurring Lockout',
            'Account Re-enable',
            'Outlook Auth Issue',
            'MFA Reset',
            'Password Expiration',
            'Temporary Account',
            'Security Investigation'
        ],
        'Step 1': [
            'Verify user identity through company app/phone system',
            'Verify user identity through company app/phone system',
            'Check Active Directory for account lockout status',
            'Check Active Directory for account lockout status',
            'Analyze lockout source using LockoutStatus.exe tool',
            'Check Active Directory for account status',
            'Diagnose Outlook authentication issue',
            'Verify user identity through alternative methods',
            'Verify password expiration in Active Directory',
            'Verify contractor authorization and requirements',
            'Analyze failed login attempt logs'
        ],
        'Step 2': [
            'Access Active Directory Users and Computers (ADUC)',
            'Access Active Directory Users and Computers (ADUC)',
            'Verify lockout was due to failed login attempts',
            'Verify lockout was due to failed login attempts',
            'Identify multiple lockout sources across domain controllers',
            'Confirm account was disabled (likely by mistake)',
            'Identify cached credential problem',
            'Access Azure AD admin center',
            'Confirm user cannot change password remotely',
            'Create temporary AD account with limited permissions',
            'Identify source IP addresses and locations'
        ],
        'Step 3': [
            'Locate user account: @username',
            'Locate user account: @username',
            'Use ADUC to unlock user account',
            'Use ADUC to unlock user account',
            'Check for cached credentials on user devices',
            'Verify user identity and authorization',
            'Clear Outlook credential cache',
            'Disable current MFA registration for user',
            'Reset password using administrative privileges',
            'Assign to appropriate security groups',
            'Verify legitimate user access patterns'
        ],
        'Step 4': [
            'Reset password using "Reset Password" function',
            'Reset password using "Reset Password" function',
            'Reset failed login counter to zero',
            'Reset failed login counter to zero',
            'Clear all cached credentials from devices',
            'Re-enable account in Active Directory Users and Computers',
            'Reset user Office 365 password',
            'Generate new MFA setup QR code',
            'Set new password with complexity requirements',
            'Set account expiration date (30 days)',
            'Implement additional security measures'
        ],
        'Step 5': [
            'Set temporary password with complexity requirements',
            'Set temporary password with complexity requirements',
            'Verify account is now accessible',
            'Verify account is now accessible',
            'Reset user password to clear cached bad passwords',
            'Verify all group memberships are intact',
            'Reconfigure Outlook with fresh credentials',
            'Provide user with new MFA setup instructions',
            'Provide user with new password via secure method',
            'Configure password policy compliance',
            'Notify user of security incident'
        ],
        'KB Article': [
            'KB_Password_Reset',
            'KB_Password_Reset',
            'KB_Password_Reset',
            'KB_Password_Reset',
            'KB_Password_Reset',
            'KB_Account_Enable',
            'KB_MFA_Reset',
            'KB_MFA_Reset',
            'KB_Password_Reset',
            'KB_Temp_Account',
            'KB_Security_Check'
        ]
    }
    
    playbook_df = pd.DataFrame(playbook_data)
    playbook_df.to_excel(writer, sheet_name='Solution Playbook', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Solution Playbook']
    format_playbook_sheet(workbook, worksheet)

def create_agent_performance_sheet(writer):
    """Create agent performance analysis sheet"""
    
    performance_data = {
        'Agent Name': ['Azola Xabadiya', 'Keawin Koesnel', 'System Admin'],
        'Total Tickets': [4, 6, 3],
        'High Priority Tickets': [1, 1, 0],
        'Medium Priority Tickets': [3, 4, 3],
        'Low Priority Tickets': [0, 1, 0],
        'Resolved Tickets': [4, 6, 3],
        'Resolution Rate': ['100%', '100%', '100%'],
        'Average Resolution Time': ['Same Day', 'Same Day', 'Same Day'],
        'Specialties': [
            'Account Management, Security Incidents',
            'Outlook Issues, MFA, Temporary Accounts',
            'Password Resets, Account Unlocks'
        ],
        'Areas for Improvement': [
            'None - Excellent performance',
            'None - Excellent performance',
            'None - Excellent performance'
        ]
    }
    
    performance_df = pd.DataFrame(performance_data)
    performance_df.to_excel(writer, sheet_name='Agent Performance', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Agent Performance']
    format_knowledge_sheet(workbook, worksheet, 'Agent Performance Analysis')

def create_common_issues_sheet(writer):
    """Create common issues and solutions sheet"""
    
    issues_data = {
        'Common Issue': [
            'User forgot password',
            'Account locked after failed attempts',
            'Recurring account lockouts',
            'Account disabled unexpectedly',
            'Outlook authentication prompts',
            'MFA device lost/stolen',
            'Password expired',
            'Contractor needs temporary access',
            'Suspicious login attempts detected'
        ],
        'User Symptoms': [
            'Cannot log in, password not working',
            'Account locked message, cannot access systems',
            'Account locks repeatedly even with correct password',
            'Login denied, account may be disabled',
            'Outlook keeps asking for password',
            'Cannot access MFA-protected services',
            'Password expired message, cannot change remotely',
            'New contractor needs system access',
            'User notified of failed login attempts'
        ],
        'Quick Diagnosis': [
            'Check if password reset is needed',
            'Verify account lockout status in AD',
            'Use LockoutStatus.exe to find source',
            'Check account status in Active Directory',
            'Check Office 365 credential cache',
            'Verify MFA device registration',
            'Check password expiration date',
            'Verify contractor authorization',
            'Review login attempt logs'
        ],
        'Standard Solution': [
            'Reset password via ADUC, provide temporary password',
            'Unlock account in ADUC, reset failed login counter',
            'Clear cached credentials from all devices',
            'Re-enable account if authorized, document reason',
            'Clear credential cache, reset Office 365 password',
            'Disable old MFA, set up new device',
            'Reset password with admin privileges',
            'Create temporary account with limited permissions',
            'Implement security measures, notify user'
        ],
        'Prevention Tips': [
            'Send password expiration reminders',
            'Educate users on correct password entry',
            'Regular credential cache maintenance',
            'Review account disablement policies',
            'Regular Office 365 credential refresh',
            'Encourage backup MFA device setup',
            'Proactive password expiration notifications',
            'Standardized contractor onboarding',
            'Enhanced security monitoring'
        ]
    }
    
    issues_df = pd.DataFrame(issues_data)
    issues_df.to_excel(writer, sheet_name='Common Issues', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Common Issues']
    format_knowledge_sheet(workbook, worksheet, 'Common Issues & Solutions')

def create_escalation_procedures_sheet(writer):
    """Create escalation procedures sheet"""
    
    escalation_data = {
        'Issue Type': [
            'High Priority Security Incident',
            'Recurring Account Lockouts',
            'Multiple User Account Issues',
            'System-wide Authentication Problems',
            'Contractor Access Violations',
            'Suspicious Login Patterns',
            'Account Disablement Disputes',
            'MFA System Failures',
            'Password Policy Violations'
        ],
        'When to Escalate': [
            'Immediate - Security breach suspected',
            'After 2 failed resolution attempts',
            'More than 5 users affected simultaneously',
            'Authentication system down',
            'Unauthorized access attempts',
            'Multiple failed logins from unknown locations',
            'User disputes account disablement reason',
            'MFA system not responding',
            'Repeated password policy violations'
        ],
        'Escalation Level 1': [
            'IT Security Team',
            'Senior IT Support',
            'IT Manager',
            'System Administrator',
            'IT Security Team',
            'IT Security Team',
            'IT Manager',
            'System Administrator',
            'IT Manager'
        ],
        'Escalation Level 2': [
            'CISO',
            'IT Director',
            'IT Director',
            'IT Director',
            'CISO',
            'CISO',
            'HR Department',
            'IT Director',
            'HR Department'
        ],
        'Documentation Required': [
            'Security incident report, log files',
            'Resolution attempts, user impact',
            'User list, affected systems',
            'System status, error logs',
            'Access logs, authorization documents',
            'Login attempt logs, IP addresses',
            'Account history, disablement reason',
            'MFA system logs, error messages',
            'Policy violations, user history'
        ]
    }
    
    escalation_df = pd.DataFrame(escalation_data)
    escalation_df.to_excel(writer, sheet_name='Escalation Procedures', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Escalation Procedures']
    format_knowledge_sheet(workbook, worksheet, 'Escalation Procedures')

def create_week1_summary_sheet(writer):
    """Create Week 1 summary sheet"""
    
    summary_data = {
        'Metric': [
            'Total Tickets Handled',
            'Tickets Resolved',
            'Resolution Rate',
            'Average Resolution Time',
            'Most Common Issue Type',
            'Highest Priority Issues',
            'Agent Performance Rating',
            'User Satisfaction',
            'Knowledge Base Articles Created',
            'Process Improvements Identified'
        ],
        'Value': [
            '13 tickets',
            '13 tickets',
            '100%',
            'Same Day',
            'Password Reset (4 tickets)',
            '2 High Priority tickets',
            'Excellent (100% resolution rate)',
            'High (all issues resolved quickly)',
            '6 KB articles',
            '5 prevention strategies identified'
        ],
        'Notes': [
            'All tickets from Week 1 successfully processed',
            'No outstanding or unresolved tickets',
            'Perfect resolution rate achieved',
            'All tickets resolved within same business day',
            'Password-related issues most frequent',
            'MFA device lost and security incidents',
            'All agents performed exceptionally well',
            'Users received prompt and effective support',
            'Comprehensive knowledge base established',
            'Proactive measures identified for common issues'
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Week 1 Summary', index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets['Week 1 Summary']
    format_knowledge_sheet(workbook, worksheet, 'Week 1 Performance Summary')

def format_knowledge_sheet(workbook, worksheet, title):
    """Format knowledge base sheets"""
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")  # Sea Green
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Apply formatting to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Set row heights
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 50
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

def format_playbook_sheet(workbook, worksheet):
    """Format the solution playbook sheet specifically"""
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=9)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set specific column widths for playbook
    column_widths = {
        'A': 20,  # Issue Category
        'B': 50,  # Step 1
        'C': 50,  # Step 2
        'D': 50,  # Step 3
        'E': 50,  # Step 4
        'F': 50,  # Step 5
        'G': 20   # KB Article
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Apply formatting to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Set row heights
    for row in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row].height = 60
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

def main():
    print("📚 Week 1 Knowledge Base Creator")
    print("================================")
    print()
    
    # Create Excel file
    filename = create_week1_knowledge_base()
    
    print()
    print("📈 Knowledge Base Summary:")
    print(f"   📄 File created: {filename}")
    print(f"   📊 Sheets included: 6 comprehensive knowledge base sections")
    print(f"   🎨 Formatting: Professional styling with sea green headers")
    
    print()
    print("✅ Week 1 Knowledge Base created successfully!")
    print(f"📁 File location: {os.path.abspath(filename)}")
    print()
    print("📋 The Knowledge Base includes:")
    print("   • Incident Patterns Analysis")
    print("   • Solution Playbook with step-by-step procedures")
    print("   • Agent Performance Analysis")
    print("   • Common Issues & Solutions")
    print("   • Escalation Procedures")
    print("   • Week 1 Performance Summary")
    print()
    print("🎯 Key Insights from Week 1:")
    print("   • 100% resolution rate achieved")
    print("   • Password-related issues most common (4 tickets)")
    print("   • All tickets resolved same day")
    print("   • 6 KB articles created for future reference")
    print("   • 5 prevention strategies identified")

if __name__ == "__main__":
    main()
